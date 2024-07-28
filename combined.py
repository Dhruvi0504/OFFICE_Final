from tkinter import *
from tkinter import messagebox, ttk
from tkcalendar import DateEntry
import openpyxl
import os

# Import dropdown data
from TypeOfMachines import type_of_machine_options
from subpart1_data import subpart1_options
from subpart2_data import subpart2_options
from subpart3_data import subpart3_options

class EntrySelectionClass:
    def __init__(self, root, return_to_dashboard):
        self.root = root
        self.return_to_dashboard_func = return_to_dashboard

        # Configure root window
        self.root.geometry("1200x700")
        self.root.title("Entry Selection")

        # Frame to contain all widgets
        self.frame = Frame(self.root, bd=2, relief=RIDGE, bg="AntiqueWhite")
        self.frame.place(x=0, y=0, relwidth=1, relheight=1)

        # Selection Label
        Label(self.frame, text="Select Entry Type:", font=("Arial", 14), bg="AntiqueWhite", fg="black").place(x=50, y=50)

        # Dropdown for selection
        self.entry_type_var = StringVar()
        self.entry_type_dropdown = ttk.Combobox(self.frame, textvariable=self.entry_type_var,
                                                values=["Purchase", "Sales"], font=("Arial", 14))
        self.entry_type_dropdown.place(x=220, y=50)
        self.entry_type_dropdown.current(0)
        self.entry_type_dropdown.bind("<<ComboboxSelected>>", self.load_entry_form)

        # Frame to contain entry fields
        self.entry_frame = Frame(self.frame, bg="AntiqueWhite")
        self.entry_frame.place(x=50, y=100, width=1100, height=600)

    def load_entry_form(self, event=None):
        for widget in self.entry_frame.winfo_children():
            widget.destroy()

        entry_type = self.entry_type_var.get()

        if entry_type == "Purchase":
            self.create_purchase_fields()
        elif entry_type == "Sales":
            self.create_sales_fields()

    def create_purchase_fields(self):
        # Invoice Number
        Label(self.entry_frame, text="Invoice Number:", **self.label_style()).place(x=50, y=60)
        self.invoice_entry = self.create_entry(self.entry_frame)
        self.invoice_entry.place(x=220, y=60)

        # Date
        Label(self.entry_frame, text="Date:", **self.label_style()).place(x=50, y=110)
        self.date_entry = DateEntry(self.entry_frame, **self.entry_style())
        self.date_entry.place(x=220, y=110)

        # Supplier Name
        Label(self.entry_frame, text="Supplier Name:", **self.label_style()).place(x=50, y=160)
        self.supplier_name_entry = self.create_entry(self.entry_frame)
        self.supplier_name_entry.place(x=220, y=160)

        # Delivery Location
        Label(self.entry_frame, text="Delivery Location:", **self.label_style()).place(x=50, y=210)
        self.delivery_location_var = StringVar(value="Classic Hitachi")
        self.delivery_location_entry = Entry(self.entry_frame, textvariable=self.delivery_location_var,
                                             **self.entry_style())
        self.delivery_location_entry.place(x=220, y=210)

        # Type of Machine (dropdown)
        Label(self.entry_frame, text="Type of Machine:", **self.label_style()).place(x=50, y=260)
        self.machine_type_var = StringVar()
        self.machine_type_dropdown = ttk.Combobox(self.entry_frame, textvariable=self.machine_type_var,
                                                  values=type_of_machine_options, font=("Arial", 12))
        self.machine_type_dropdown.place(x=220, y=260)
        self.machine_type_dropdown.current(0)

        # SubPart 1 beside Type of Machine (increased space)
        Label(self.entry_frame, text="SubPart 1:", **self.label_style()).place(x=400, y=260)  # Adjusted x for spacing
        self.dropdown1_var = StringVar()
        self.dropdown1 = ttk.Combobox(self.entry_frame, textvariable=self.dropdown1_var,
                                      values=subpart1_options, font=("Arial", 12))
        self.dropdown1.place(x=475, y=260)  # Adjusted x for spacing
        self.dropdown1.current(0)

        # SubPart 2
        Label(self.entry_frame, text="SubPart 2:", **self.label_style()).place(x=50, y=310)
        self.dropdown2_var = StringVar()
        self.dropdown2 = ttk.Combobox(self.entry_frame, textvariable=self.dropdown2_var,
                                      values=subpart2_options, font=("Arial", 12))
        self.dropdown2.place(x=220, y=310)
        self.dropdown2.current(0)

        # SubPart 3 beside SubPart 2
        Label(self.entry_frame, text="SubPart 3:", **self.label_style()).place(x=400, y=310)  # Adjusted x for spacing
        self.dropdown3_var = StringVar()
        self.dropdown3 = ttk.Combobox(self.entry_frame, textvariable=self.dropdown3_var,
                                      values=subpart3_options, font=("Arial", 12))
        self.dropdown3.place(x=475, y=310)  # Adjusted x for spacing
        self.dropdown3.current(0)

        # Quantity
        Label(self.entry_frame, text="Quantity:", **self.label_style()).place(x=50, y=360)
        self.quantity_entry = self.create_entry(self.entry_frame)
        self.quantity_entry.place(x=220, y=360)

        # Unit Price
        Label(self.entry_frame, text="Unit Price:", **self.label_style()).place(x=50, y=410)
        self.price_entry = self.create_entry(self.entry_frame)
        self.price_entry.place(x=220, y=410)

        # Total (displayed as a Label)
        Label(self.entry_frame, text="Total:", **self.label_style()).place(x=50, y=460)
        self.total_display = Label(self.entry_frame, text="0", font=("Arial", 14), bg="AntiqueWhite", fg="black")
        self.total_display.place(x=220, y=460)

        # Automatically calculate total when quantity or price changes
        self.quantity_entry.bind("<KeyRelease>", self.update_total)
        self.price_entry.bind("<KeyRelease>", self.update_total)

        # Save button
        btn_save = Button(self.entry_frame, text="Save", command=self.export_to_excel, font=("Arial", 14),
                          bg="lightgray", fg="black", bd=3, cursor="hand2")
        btn_save.place(relx=1.0, rely=0.0, anchor=NE, x=-10, y=10)

        # Back to Dashboard button
        btn_back = Button(self.entry_frame, text="Back to Dashboard", command=self.return_to_dashboard,
                          font=("Arial", 14),
                          bg="lightgray", fg="black", bd=3, cursor="hand2")
        btn_back.place(relx=1.0, rely=0.0, anchor=NE, x=-10, y=60)

        # Clear button
        btn_clear = Button(self.entry_frame, text="Clear", command=self.clear_fields, font=("Arial", 14),
                           bg="lightgray", fg="black", bd=3, cursor="hand2")
        btn_clear.place(relx=1.0, rely=0.0, anchor=NE, x=-10, y=110)

    def create_sales_fields(self):
        # Invoice Number
        Label(self.entry_frame, text="Invoice Number:", **self.label_style()).place(x=50, y=60)
        self.invoice_entry = self.create_entry(self.entry_frame)
        self.invoice_entry.place(x=220, y=60)

        # Date
        Label(self.entry_frame, text="Date:", **self.label_style()).place(x=50, y=110)
        self.date_entry = DateEntry(self.entry_frame, **self.entry_style())
        self.date_entry.place(x=220, y=110)

        # Customer Name
        Label(self.entry_frame, text="Customer Name:", **self.label_style()).place(x=50, y=160)
        self.customer_name_entry = self.create_entry(self.entry_frame)
        self.customer_name_entry.place(x=220, y=160)

        # Delivery Location
        Label(self.entry_frame, text="Delivery Location:", **self.label_style()).place(x=50, y=210)
        self.delivery_location_var = StringVar()
        self.delivery_location_entry = Entry(self.entry_frame, textvariable=self.delivery_location_var,
                                             **self.entry_style())
        self.delivery_location_entry.place(x=220, y=210)

        # Type of Machine (dropdown)
        Label(self.entry_frame, text="Type of Machine:", **self.label_style()).place(x=50, y=260)
        self.machine_type_var = StringVar()
        self.machine_type_dropdown = ttk.Combobox(self.entry_frame, textvariable=self.machine_type_var,
                                                  values=type_of_machine_options, font=("Arial", 12))
        self.machine_type_dropdown.place(x=220, y=260)
        self.machine_type_dropdown.current(0)

        # SubPart 1 beside Type of Machine (shifted slightly right)
        Label(self.entry_frame, text="SubPart 1:", **self.label_style()).place(x=410, y=260)  # Adjusted x for spacing
        self.dropdown1_var = StringVar()
        self.dropdown1 = ttk.Combobox(self.entry_frame, textvariable=self.dropdown1_var,
                                      values=subpart1_options, font=("Arial", 12))
        self.dropdown1.place(x=485, y=260)  # Adjusted x for spacing
        self.dropdown1.current(0)

        # SubPart 2
        Label(self.entry_frame, text="SubPart 2:", **self.label_style()).place(x=50, y=310)
        self.dropdown2_var = StringVar()
        self.dropdown2 = ttk.Combobox(self.entry_frame, textvariable=self.dropdown2_var,
                                      values=subpart2_options, font=("Arial", 12))
        self.dropdown2.place(x=220, y=310)
        self.dropdown2.current(0)

        # SubPart 3 beside SubPart 2 (shifted slightly right)
        Label(self.entry_frame, text="SubPart 3:", **self.label_style()).place(x=410, y=310)  # Adjusted x for spacing
        self.dropdown3_var = StringVar()
        self.dropdown3 = ttk.Combobox(self.entry_frame, textvariable=self.dropdown3_var,
                                      values=subpart3_options, font=("Arial", 12))
        self.dropdown3.place(x=485, y=310)  # Adjusted x for spacing
        self.dropdown3.current(0)

        # Quantity
        Label(self.entry_frame, text="Quantity:", **self.label_style()).place(x=50, y=360)
        self.quantity_entry = self.create_entry(self.entry_frame)
        self.quantity_entry.place(x=220, y=360)

        # Save button
        btn_save = Button(self.entry_frame, text="Save", command=self.export_to_excel, font=("Arial", 14),
                          bg="lightgray", fg="black", bd=3, cursor="hand2")
        btn_save.place(relx=1.0, rely=0.0, anchor=NE, x=-10, y=10)

        # Back to Dashboard button
        btn_back = Button(self.entry_frame, text="Back to Dashboard", command=self.return_to_dashboard,
                          font=("Arial", 14),
                          bg="lightgray", fg="black", bd=3, cursor="hand2")
        btn_back.place(relx=1.0, rely=0.0, anchor=NE, x=-10, y=60)

        # Clear button
        btn_clear = Button(self.entry_frame, text="Clear", command=self.clear_fields, font=("Arial", 14),
                           bg="lightgray", fg="black", bd=3, cursor="hand2")
        btn_clear.place(relx=1.0, rely=0.0, anchor=NE, x=-10, y=110)

    def export_to_excel(self):
        try:
            if not self.invoice_entry.get():
                messagebox.showerror("Error", "Please enter Invoice Number")
                return

            file_path = "Inventory_Data.xlsx"

            if not os.path.exists(file_path):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Inventory Data"
                sheet.append([
                    "Entry Type", "Invoice Number", "Date", "Supplier/Customer Name", "Delivery Location",
                    "Type of Machine", "SubPart 1", "SubPart 2", "SubPart 3",
                    "Quantity", "Unit Price", "Total"
                ])
                workbook.save(file_path)

            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            entry_data = [
                self.entry_type_var.get(),  # Save Entry Type
                self.invoice_entry.get(),
                self.date_entry.get_date().strftime("%Y-%m-%d"),
                self.supplier_name_entry.get() if self.entry_type_var.get() == "Purchase" else self.customer_name_entry.get(),
                self.delivery_location_entry.get(),
                self.machine_type_dropdown.get(),
                self.dropdown1.get(),
                self.dropdown2.get(),
                self.dropdown3.get(),
                self.quantity_entry.get()
            ]

            if self.entry_type_var.get() == "Purchase":
                entry_data.extend([
                    self.price_entry.get(),  # Unit Price
                    self.total_label.cget("text")  # Total
                ])
            else:
                entry_data.extend(["-", "-"])  # Unit Price and Total as 0 for Sales

            sheet.append(entry_data)
            workbook.save(file_path)
            messagebox.showinfo("Success", "Data saved successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

    def label_style(self):
        return {"font": ("Arial", 14), "bg": "AntiqueWhite", "fg": "black"}

    def entry_style(self):
        return {"font": ("Arial", 12), "width": 30}

    def create_entry(self, parent):
        return Entry(parent, font=("Arial", 12), width=30)

    def update_total(self, event=None):
        try:
            quantity = float(self.quantity_entry.get() or 0)
            unit_price = float(self.price_entry.get() or 0)
            total = quantity * unit_price
            self.total_display.config(text=f"{total:.2f}")
        except ValueError:
            self.total_display.config(text="0")

    def export_to_excel(self):
        try:
            if not self.invoice_entry.get():
                messagebox.showerror("Error", "Please enter Invoice Number")
                return

            file_path = "Inventory_Data.xlsx"

            if not os.path.exists(file_path):
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Inventory Data"
                sheet.append([
                    "Entry Type", "Invoice Number", "Date", "Supplier/Customer Name", "Delivery Location",
                    "Type of Machine", "SubPart 1", "SubPart 2", "SubPart 3",
                    "Quantity", "Unit Price", "Total"
                ])
                workbook.save(file_path)

            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            entry_data = [
                self.entry_type_var.get(),  # Save Entry Type
                self.invoice_entry.get(),
                self.date_entry.get_date().strftime("%Y-%m-%d"),
                self.supplier_name_entry.get() if self.entry_type_var.get() == "Purchase" else self.customer_name_entry.get(),
                self.delivery_location_entry.get(),
                self.machine_type_dropdown.get(),
                self.dropdown1.get(),
                self.dropdown2.get(),
                self.dropdown3.get(),
                self.quantity_entry.get()
            ]

            if self.entry_type_var.get() == "Purchase":
                entry_data.extend([
                    self.price_entry.get(),  # Unit Price
                    self.total_display.cget("text")  # Total
                ])
            else:
                entry_data.extend([0, 0])  # Unit Price and Total as 0 for Sales

            sheet.append(entry_data)
            workbook.save(file_path)
            messagebox.showinfo("Success", "Data saved successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

    def clear_fields(self):
        self.invoice_entry.delete(0, END)
        self.date_entry.set_date("")
        if hasattr(self, 'supplier_name_entry'):
            self.supplier_name_entry.delete(0, END)
        if hasattr(self, 'customer_name_entry'):
            self.customer_name_entry.delete(0, END)
        self.delivery_location_entry.delete(0, END)
        self.machine_type_dropdown.set("")
        self.dropdown1.set("")
        self.dropdown2.set("")
        self.dropdown3.set("")
        self.quantity_entry.delete(0, END)
        self.price_entry.delete(0, END)
        self.total_label.config(text="")

    def return_to_dashboard(self):
        self.frame.destroy()
        self.return_to_dashboard_func()

if __name__ == "__main__":
    root = Tk()
    def go_to_dashboard():
        print("Redirecting to dashboard...")
        # Here you can add the code to open the dashboard window
        # For example, you might destroy the current window and create a new one

    app = EntrySelectionClass(root, go_to_dashboard)
    root.mainloop()
